"""
สร้าง findings-log.xlsx template สำหรับ KAS Client Knowledge Base.
ใช้ openpyxl เพื่อ: header formatting, freeze panes, autofilter, ตัวอย่างข้อมูล

Usage: python create-findings-template.py <output_path>
"""
import sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def create_findings_template(output_path: str):
    wb = openpyxl.Workbook()

    # ============================================================
    # Sheet 1: ข้อตรวจพบ (Findings Log)
    # ============================================================
    ws1 = wb.active
    ws1.title = "Findings Log"

    # Styling
    header_font = Font(name='Angsana New', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    data_font = Font(name='Angsana New', size=11)

    headers = [
        'ปี', 'เล่มที่/รายงาน', 'วันที่ออกรายงาน',
        'Finding No.', 'ชื่อข้อตรวจพบ',
        'กระบวนการ', 'ระดับความเสี่ยง\n(H/M/L)',
        'ประเภท\n(Key/Non-Key)', 'สถานะ\n(Open/Closed)',
        'วันที่ closed', 'คำอธิบายโดยสรุป',
        'หน่วยงานรับผิดชอบ', 'หมายเหตุ'
    ]

    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Column widths
    col_widths = [6, 20, 14, 10, 30, 20, 12, 12, 12, 14, 40, 20, 20]
    for i, w in enumerate(col_widths, 1):
        ws1.column_dimensions[get_column_letter(i)].width = w

    # Example data rows (3 rows to show format)
    examples = [
        ['2567', 'IA Report Q1/2567', '15/01/2567', 'F-001',
         'การอนุมัติการจ่ายเงินไม่เป็นไปตามอำนาจดำเนินการ', 'Disbursement',
         'H', 'Key', 'Open', '',
         'พบการจ่ายเงินเกินอำนาจอนุมัติ 5 รายการ รวม 2.5 ลบ.', 'ฝ่ายบัญชี', ''],
        ['2567', 'IA Report Q1/2567', '15/01/2567', 'F-002',
         'การตรวจนับสินค้าคงคลังไม่ครอบคลุมทุกรายการ', 'Inventory',
         'M', 'Non-Key', 'Closed', '28/02/2567',
         'ตรวจนับได้ 85% ของรายการ คงเหลือ 15% ที่ไม่ได้รับการตรวจนับ',
         'ฝ่ายคลังสินค้า', 'closed หลัง follow-up Q2'],
        ['2566', 'IA Report 2566', '05/08/2566', 'F-001',
         'ไม่มีการสอบทานการเข้าถึงระบบ ERP โดย IT', 'IT General Control',
         'H', 'Key', 'Closed', '15/12/2566',
         'user admin ไม่เคยถูก review ตั้งแต่ go-live ปี 2564', 'ฝ่าย IT', ''],
    ]

    for r, row_data in enumerate(examples, 2):
        for c, val in enumerate(row_data, 1):
            cell = ws1.cell(row=r, column=c, value=val)
            cell.font = data_font
            cell.alignment = Alignment(vertical='center', wrap_text=True)
            cell.border = thin_border

    # Freeze header + autofilter
    ws1.freeze_panes = 'A2'
    ws1.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(examples)+1}"

    # ============================================================
    # Sheet 2: คำแนะนำ (Instructions)
    # ============================================================
    ws2 = wb.create_sheet("วิธีใช้")
    instructions = [
        "📋 วิธีกรอก Findings Log",
        "",
        "1. แต่ละแถว = 1 ข้อตรวจพบ (1 Finding)",
        "2. Finding No. ให้ใส่เลขเรียงตามเล่มรายงาน (F-001, F-002, ...)",
        "3. สถานะเริ่มต้น = 'Open' → เมื่อแก้ไขแล้วเปลี่ยนเป็น 'Closed' พร้อมวันที่",
        "4. ระดับความเสี่ยง: H = High, M = Medium, L = Low",
        "5. ประเภท: Key = ประเด็นสำคัญที่มีนัยสำคัญ, Non-Key = ข้อเสนอแนะทั่วไป",
        "6. ห้ามลบแถว — ถ้า closed แล้ว ให้เปลี่ยนสถานะ ไม่ต้องลบ",
        "7. เมื่อออกรายงานใหม่ → เพิ่มแถวต่อท้าย",
        "8. 🤖 Hermes อ่านไฟล์นี้เพื่อสร้าง dashboard ติดตามสถานะอัตโนมัติ",
    ]
    for i, row in enumerate(instructions, 1):
        ws2.cell(row=i, column=1, value=row).font = Font(
            name='Angsana New', size=12, bold=(i == 1)
        )
    ws2.column_dimensions['A'].width = 80

    wb.save(output_path)
    print(f"✅ Created: {output_path}")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python create-findings-template.py <output_path>")
        sys.exit(1)
    create_findings_template(sys.argv[1])
